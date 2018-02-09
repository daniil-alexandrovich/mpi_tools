'''
PortfolioUpdater manipulates portfolio data and writes it, properly
formatted for Stylus, to a .xlsx file.

Uses a pandas DataFrame object to generate portfolios consisting of the
following data for each underlying fund:
    
    * ID (Index) - Fund ID within DBID
    *   Label    - Referential label within portfolio
    *    DBID    - ID of parent database in Stylus
    *   Date 1   - Weight of asset at Date 1
    *   Date 2   - Weight of asset at Date 2
    *    ...

Potential Future Methods for manual adjustments:
    - new(): create new portfolio
    - add_fund(portfolio, id, label, dbid): add fund to portfolio
    - add_date(portfolio, date): add date
    - change_weight(portfolio, id, date, weight): change a weight

Created on 1/9/2018 by Daniil Feoktistov for Markov Processes International, Inc.
'''

# A class for generating and editing Stylus Pro portfolios

import  pandas  as pd
import openpyxl as xl
import    os
import    re
import    sys
import datetime

class PortfolioUpdater:
    def load(self, filename, sheetname, path=None, stylus_formatted=False):
        """Loads an Excel worksheet in as a portfolio object.
        
        Retrieves data from an Excel worksheet stored path/filename:sheet and
        returns it as a DataFrame. If input data is not Stylus-formatted, it
        must instead be formatted as follows:
            
                        A         B       C       D     E      ...
                   1   ID      Label    DBID  <date> <date>
                   2 FOUSA1  MStarFund  MfX     10   45.678
                   3 012345  eVestFund  eVa            0     
                   ...
        
        Args:
            filename: Target file name
            sheet: Target sheet name
            path: Filepath of Excel sheet if not cwd.
            stylus_formatted: If the worksheet is formatted as a Stylus
                portfolio, it is parsed and read from accordingly.
        
        Returns:
          A DataFrame with the following column order:
              
              0: (Index) ID
              1. Label
              2. DBID
              3+. Date (datetime.datetime object)
        """
        filename = os.path.join(path, filename) if path else filename
        sheet = xl.load_workbook(filename)[sheetname]
        if stylus_formatted:
            # Sheet metadata (MPI_ASSETIDRANGE, MPI_REBALANCE, etc.) is stored
            # and passed through to the output file.
            metadata, cellrange = self.get_metadata(sheet)
            portfolio = self.get_portfolio(sheet, cellrange)
            df = pd.DataFrame(portfolio)
            # Names fields by first row and drops that row.
            df.rename(columns = df.iloc[0], inplace=True)
            df.drop(df.index[0], inplace=True)
            df.set_index('ID', inplace=True)
            return df, metadata
        else:
            # Case where no metadata is provided.
            if sheet[1][0].value == 'ID':
                df = pd.read_excel(filename, sheetname=sheetname, index_col=0)
                metadata = None
            # Case where metadata is provided.
            else:
                df = pd.read_excel(filename, sheetname=sheetname, skiprows=[0,1], index_col=0)
                metadata = {}
                for key, value in zip(sheet['1'], sheet['2']):
                    if key.value:
                        metadata[key.value] = value.value
            return df, metadata
            

    def add_dates(self, portfolio, additions):
        # Name parameters:
            # f(self,*, x, y...)
            # x = f(y=yval, x=xval..)
            
        '''Outer merges two portfolio objects while retaining indexing by ID.
        
        The outer merge process creates rows in <portfolio> for entries in
        <additions> that are not shared. If a fund exists in one portfolio and
        not the other, its values for dates it has not been assigned weights
        are set to zero.
        
        After this process, date columns are resorted in order to allow for
        insertion of weight dates in addition to apppending.
        
        Args:
            portfolio: DataFrame object containing portfolio into which
                additional data is merged.
            additions: DataFrame object containing portfolio from which
                additional data is merged.
        
        Returns:
            DataFrame containing union of rows and columns from 'portfolio' and
            'additions'. Date columns are sorted
        '''
        portfolio.reset_index(inplace=True)
        portfolio = portfolio.merge(additions.reset_index(), how='outer')
        portfolio.set_index('ID', inplace=True)
        # Dates are separated out temporarily in order to sort them while
        # retaining order of the first few columns.
        dates_sorted = portfolio.iloc[:,2:].sort_index(axis=1)
        return pd.concat([portfolio.iloc[:,:2], dates_sorted], axis=1)
        
    
    def write(self, portfolio, metadata, filename, sheet):
        '''Writes an Advanced Portfolio to a Stylus-formatted Excel sheet.
        
        The output worksheet is written the first two rows set aside for
        metadata and portfolio contents from row 5 onwards. Dates occupy row 4.
        '''
#        metadata = self.update_metadata(portfolio, metadata, joined_metadata)
        out_book = xl.load_workbook(filename=filename)
        output = out_book[sheet]
        # Metadata is written to first two rows.        
        for col, key in enumerate(metadata.keys()):
            output.cell(row=1, column=col+1).value = key
            output.cell(row=2, column=col+1).value = metadata[key]
        # Dates from header row are reformatted from datetime and written.
        for col, header in enumerate(portfolio.axes[1]):
            if type(header) is datetime.datetime:
                output.cell(row=4, column=col+2).value = header.date()
        # Body of portfolio data is parsed from DataFrame and written.
        for row, asset in enumerate(zip(portfolio.iterrows())):
            for id, data in asset:
                output.cell(row=row+5, column=1).value = id
                for col, point in enumerate(data):
                    output.cell(row=row+5, column=col+2).value = point
        out_book.save(filename)
        
    def get_metadata(self, sheet):
        '''Retrieves portfolio metadata from a Stylus-formatted sheet.
        
        Returns a dictionary containing metadata and a string representing
        the cell range of portfolio data in the worksheet.
        '''
        metadata = {}
        # Rows 1 and 2 contain metadata parameter and value, respectively
        for key, value in zip(sheet['1'], sheet['2']):
            if key.value:
                metadata[key.value] = value.value
        # The range of cells to be red is determined from the row range of
        # MPI_LABELRANGE and the column range of MPI_PORTFOLIODATERANGE.
        rowrange = metadata['MPI_LABELRANGE'].split(":")
        colrange = metadata['MPI_PORTFOLIODATERANGE'].split(":")
        cellrange = ['','']
        cellrange[0] = 'A'+str(int(rowrange[0][1:])-1)
        # A regex is used to to isolate the row number from the cell identifier.
        cellrange[1] = re.split('(\d+)',colrange[1])[0]+rowrange[1][1:]
        # The remaining range is parsed into a list-of-lists, which is then
        # passed into a pandas DataFrame.
        return metadata, cellrange
    
    def get_portfolio(self, sheet, cellrange):
        '''Returns a nested list representing portfolio data fom a Stylus-
        formatted Excel worksheet.'''
        portfolio = []
        rows = sheet[cellrange[0]:cellrange[1]]
        # Header row is loaded in and updated to include field names that are
        # not included in the worksheet.
        header_row = [cell.value for cell in rows[0]]
        header_row[:3] = ['ID', 'Label', 'DBID']
        portfolio.append(header_row)     
        # Remaining data can then be loaded in without adjustment.
        for row in rows[1:]:
            values = [cell.value for cell in row]
            portfolio.append(values)
        return portfolio
    
    def update_metadata(self, portfolio, metadata, joined_metadata=None):
        '''Updates existing metadata to reflect a portfolio's new cell ranges
        
        If None is passed, new metadata for the portfolio is generated.
        '''
        if not metadata:
            metadata = {}
        if 'MPI_Rebalance' not in metadata:
            metadata['MPI_Rebalance'] = 'Monthly'
        metadata['MPI_PORTFOLIOTYPE'] = 'Advanced'
        # Existing metadata is overwritten in order to accomodate for added rows or columns.
        last_row = len(portfolio) + 4
        last_col = xl.utils.get_column_letter(len(portfolio.axes[1])+1)
        metadata['MPI_ASSETIDRANGE'] = 'A5:A'+str(last_row)
        metadata['MPI_LABELRANGE'] = 'B5:B'+str(last_row)
        metadata['MPI_ASSETDBIDRANGE'] = 'C5:C'+str(last_row)
        metadata['MPI_PORTFOLIODATERANGE'] = 'D4:'+last_col+'4'
        # Metadata from portfolio second portfolio is added to existing metadata.
        if joined_metadata:
            for key in joined_metadata:
                if key not in metadata:
                    metadata[key] = joined_metadata[key]
        return metadata
    
    def run(self, file, sheet, stylus_formatted, out_file, out_sheet, existing_file=None, existing_sheet=None, existing_stylus_formatted=None):
        print('\nLibraries loaded.')
        data, meta = self.load(file, sheet, stylus_formatted=stylus_formatted)
        print('Data imported from '+file+'.')
        if existing_file:
            existing_data, existing_meta = self.load(existing_file, existing_sheet, stylus_formatted=existing_stylus_formatted)
            print('Data imported from '+existing_file+'.')
            data = self.add_dates(existing_data, data)
            meta = self.update_metadata(data, existing_meta, meta)
            print('Data merged.')
        else:
            meta = self.update_metadata(data, meta)
        self.write(data, meta, out_file, out_sheet)
        print('Data from exported to '+out_file+'.')    

if __name__ == '__main__':
    print()
    if len(sys.argv) not in [5,7]:
        print('Please pass either 4 or 6 parameters!')
        sys.exit(0)
    pu = PortfolioUpdater()
    if len(sys.argv) == 5:
        stylus_formatted = False if int(input('Is '+sys.argv[1]+' Stylus-formatted? (0 for No, 1 for Yes) >> ')) == 0 else True
        pu.run(sys.argv[1], sys.argv[2], stylus_formatted, sys.argv[3], sys.argv[4])
    elif len(sys.argv) == 7:
        stylus_formatted = False if int(input('Is '+sys.argv[1]+' Stylus-formatted? (0 for No, 1 for Yes) >> ')) == 0 else True
        existing_stylus_formatted = False if int(input('Is '+sys.argv[5]+' Stylus-formatted? (0 for No, 1 for Yes) >> ')) == 0 else True
        pu.run(sys.argv[1], sys.argv[2], stylus_formatted, sys.argv[3], sys.argv[4], sys.argv[5], sys.argv[6], existing_stylus_formatted)
    